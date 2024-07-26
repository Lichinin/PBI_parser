import base64
import datetime
import os

import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side


class DMM1510D:
    def __init__(self, device_ip, username, password, location):
        self.device_ip = device_ip
        self.username = username
        self.password = password
        self.location = location
        self.headers = {
            'Authorization': 'Basic ' + base64.b64encode(f'{username}:{password}'.encode('utf-8')).decode('utf-8')
        }

    def check_connection(self):
        tuner_url = f'http://{self.username}:{self.password}@{self.device_ip}'
        response = requests.get(tuner_url)
        return response.status_code

    def get_tuner_parameters(self):
        tuner_url = f'http://{self.device_ip}/cgi-bin/tuner_config.cgi'

        tuner_response = requests.get(tuner_url, headers=self.headers)

        soup = BeautifulSoup(tuner_response.text, 'html.parser')

        self.tuner_lnb = soup.find(
            'input',
            {'name': 'LnbFreq'}
        ).get('value')
        self.tuner_satellite_frequency = soup.find(
            'input',
            {'name': 'SateFreq'}
        ).get('value')
        self.tuner_symbol_rate = soup.find(
            'input',
            {'name': 'SateSr'}
        ).get('value')
        lnb_voltage_raw = soup.find('select', {'name': 'lnbVol'})
        self.tuner_lnb_voltage = (((lnb_voltage_raw.find(
            'option',
            {'selected': True})
        ).text).split('\n'))[0]
        lnb_22khz_raw = soup.find('select', {'name': 'lnb22k'})
        self.tuner_lnb_22khz = (((lnb_22khz_raw.find(
            'option',
            {'selected': True})).text
        ).split('\n'))[0]

    def get_remux_parameters(self):
        response = requests.get(
            f'http://{self.device_ip}/cgi-bin/mux_config.cgi',
            headers=self.headers
        )
        response.encoding = 'utf-8'

        soup = BeautifulSoup(response.text, 'html.parser')

        tuner_data = soup.find('div', {'id': 'source2_out_value'})
        tuner_raw = tuner_data.find_all('div', {'class': 'tree_3'})

        self.tuner_outputs = [
            div.text.strip().split('\xa0')[-1]
            for div in tuner_raw
        ]

        ci_data = soup.find('div', {'id': 'source4_out_value'})
        ci_raw = ci_data.find_all('div', {'class': 'tree_3'})

        self.ci_outputs = [
            div.text.strip().split('\xa0')[-1]
            for div in ci_raw
        ]

        ip_data = soup.find('div', {'id': 'source3_out_value'})
        ip_raw = ip_data.find_all('div', {'class': 'tree_3'})

        self.ip_outputs = [
            div.text.strip().split('\xa0')[-1]
            for div in ip_raw
        ]

    def get_decoder(self):
        response = requests.get(
            f'http://{self.device_ip}/cgi-bin/decoder_config.cgi',
            headers=self.headers
        )
        response.encoding = 'utf-8'

        soup = BeautifulSoup(response.text, 'html.parser')

        decoder_data = soup.find(
            'input',
            {'name': 'service_name'}
        )
        self.decoder = decoder_data['value']

    def get_ip_input(self):
        ip_parts = []
        response = requests.get(
            f'http://{self.device_ip}/cgi-bin/ipin.cgi',
            headers=self.headers
        )
        soup = BeautifulSoup(response.text, 'html.parser')

        for i in range(0, 4):
            input_name = f'gigabit_uni_multicast_in_address_0{i}'
            input_tag = soup.find('input', {'name': input_name})
            if input_tag:
                ip_parts.append(input_tag['value'])
            else:
                break
        ip_address = '.'.join(ip_parts)
        port = soup.find(
            'input',
            {'name': 'gigabit_uni_multicast_udp_in_port_1'}
        ).get('value')
        self.input_ip = f'{ip_address}:{port}'

    def get_ip_params(self):
        ip_type = self.check_ip_type()
        if ip_type == 'TS/IP In':
            self.get_ip_input()
        elif ip_type == 'TS/IP Out':
            self.get_ip_output()

    def get_all_parameters(self):
        if self.check_connection() != 200:
            raise ConnectionError('Нет связи с устройством')
        self.get_tuner_parameters()
        self.get_remux_parameters()
        self.get_decoder()
        self.get_ip_input()

    def export_params_to_excel(self, name):
        output_dir = os.path.join(os.getcwd(), "excel_output")
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        file_path = os.path.join(
            output_dir,
            f"{self.location} {name} PBI {self.__class__.__name__} {datetime.date.today()}.xlsx"
        )
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
        worksheet = workbook.active
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        worksheet.column_dimensions['A'].width = 20

        worksheet['A1'] = f'Конфигурация устройства {self.device_ip} ({self.__class__.__name__}) ({self.location}) {datetime.date.today()}'
        worksheet['A1'].font = Font(bold=True)
        row = worksheet.max_row + 2

        worksheet.cell(
            row=row,
            column=1,
            value='Параметры Tuner'
        ).font = Font(bold=True)
        row = worksheet.max_row + 1

        worksheet.cell(
            row=row,
            column=1,
            value='LNB'
        ).border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner_lnb
        ).border = border
        row += 1

        worksheet.cell(
            row=row,
            column=1,
            value='Satellite Frequency'
        ).border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner_satellite_frequency
        ).border = border
        row += 1

        worksheet.cell(
            row=row,
            column=1,
            value='Symbol Rate'
        ).border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner_symbol_rate
        ).border = border
        row += 1

        worksheet.cell(
            row=row,
            column=1,
            value='LNB Voltage'
        ).border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner_lnb_voltage
        ).border = border
        row += 1

        worksheet.cell(
            row=row,
            column=1,
            value='LNB 22KHz'
        ).border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner_lnb_22khz
        ).border = border
        row += 2

        worksheet.cell(
            row=row,
            column=1,
            value='Параметры Remux'
        ).font = Font(bold=True)
        row = worksheet.max_row + 1

        worksheet.cell(
            row=row,
            column=1,
            value='Tuner Out'
        ).border = border
        worksheet.cell(
            row=row,
            column=2,
            value=", ".join(str(x) for x in self.tuner_outputs)
        ).border = border
        row += 1

        worksheet.cell(
            row=row,
            column=1,
            value='CI Out'
        ).border = border
        worksheet.cell(
            row=row,
            column=2,
            value=", ".join(str(x) for x in self.ci_outputs)
        ).border = border
        row += 1

        worksheet.cell(
            row=row,
            column=1,
            value='IP Out'
        ).border = border
        worksheet.cell(
            row=row,
            column=2,
            value=", ".join(str(x) for x in self.ip_outputs)
        ).border = border
        row += 2

        worksheet.cell(
            row=row,
            column=1,
            value='Декодер'
        ).font = Font(bold=True)
        row += 1

        worksheet.cell(
            row=row,
            column=1,
            value='Выход декодера'
        ).border = border
        worksheet.cell(
            row=row,
            column=2,
            value=str(self.decoder)
        ).border = border
        row += 2

        worksheet.cell(
            row=row,
            column=1,
            value='Мультикаст'
        ).font = Font(bold=True)
        row += 1

        worksheet.cell(
            row=row,
            column=1,
            value='IP multicast IN'
        ).border = border
        worksheet.cell(
            row=row,
            column=2,
            value=str(self.input_ip)
        ).border = border
        row += 2

        workbook.save(file_path)
