import base64
import datetime
import os

import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side


class DMM2400D:
    def __init__(self, device_ip, username, password, location):
        self.device_ip = device_ip
        self.location = location
        self.headers = {
            'Authorization': 'Basic ' + base64.b64encode(f'{username}:{password}'.encode('utf-8')).decode('utf-8')
        }

    def get_tuner1_parameters(self):
        tuner_url = f'http://{self.device_ip}/cgi-bin/tuner1.cgi'

        tuner_response = requests.get(tuner_url, headers=self.headers)

        if tuner_response.status_code == 200:
            soup = BeautifulSoup(tuner_response.text, 'html.parser')

            self.tuner1_lnb = soup.find(
                'input',
                {'name': 'LnbFreq'}
            ).get('value')
            self.tuner1_satellite_frequency = soup.find(
                'input',
                {'name': 'SateFreq'}
            ).get('value')
            self.tuner1_symbol_rate = soup.find(
                'input',
                {'name': 'SateSr'}
            ).get('value')
            lnb_voltage_raw = soup.find('select', {'name': 'lnbVol'})
            self.tuner1_lnb_voltage = (((lnb_voltage_raw.find(
                'option',
                {'selected': True})
            ).text).split('\n'))[0]
            lnb_22khz_raw = soup.find('select', {'name': 'lnb22k'})
            self.tuner1_lnb_22khz = (((lnb_22khz_raw.find(
                'option',
                {'selected': True})).text
            ).split('\n'))[0]

        else:
            print('Ошибка запроса:', tuner_response.status_code)

    def get_tuner2_parameters(self):
        tuner_url = f'http://{self.device_ip}/cgi-bin/tuner2.cgi'

        tuner_response = requests.get(tuner_url, headers=self.headers)

        if tuner_response.status_code == 200:
            soup = BeautifulSoup(tuner_response.text, 'html.parser')

            self.tuner2_lnb = soup.find(
                'input',
                {'name': 'LnbFreq'}
            ).get('value')
            self.tuner2_satellite_frequency = soup.find(
                'input',
                {'name': 'SateFreq'}
            ).get('value')
            self.tuner2_symbol_rate = soup.find(
                'input',
                {'name': 'SateSr'}
            ).get('value')
            lnb_voltage_raw = soup.find('select', {'name': 'lnbVol'})
            self.tuner2_lnb_voltage = (((lnb_voltage_raw.find(
                'option',
                {'selected': True})
            ).text).split('\n'))[0]
            lnb_22khz_raw = soup.find('select', {'name': 'lnb22k'})
            self.tuner2_lnb_22khz = (((lnb_22khz_raw.find(
                'option',
                {'selected': True})).text
            ).split('\n'))[0]

        else:
            print('Ошибка запроса:', tuner_response.status_code)

    def get_tuner3_parameters(self):
        tuner_url = f'http://{self.device_ip}/cgi-bin/tuner3.cgi'

        tuner_response = requests.get(tuner_url, headers=self.headers)

        if tuner_response.status_code == 200:
            soup = BeautifulSoup(tuner_response.text, 'html.parser')

            self.tuner3_lnb = soup.find(
                'input',
                {'name': 'LnbFreq'}
            ).get('value')
            self.tuner3_satellite_frequency = soup.find(
                'input',
                {'name': 'SateFreq'}
            ).get('value')
            self.tuner3_symbol_rate = soup.find(
                'input',
                {'name': 'SateSr'}
            ).get('value')
        else:
            print('Ошибка запроса:', tuner_response.status_code)

    def get_tuner4_parameters(self):
        tuner_url = f'http://{self.device_ip}/cgi-bin/tuner4.cgi'

        tuner_response = requests.get(tuner_url, headers=self.headers)

        if tuner_response.status_code == 200:
            soup = BeautifulSoup(tuner_response.text, 'html.parser')

            self.tuner4_lnb = soup.find(
                'input',
                {'name': 'LnbFreq'}
            ).get('value')
            self.tuner4_satellite_frequency = soup.find(
                'input',
                {'name': 'SateFreq'}
            ).get('value')
            self.tuner4_symbol_rate = soup.find(
                'input',
                {'name': 'SateSr'}
            ).get('value')
        else:
            print('Ошибка запроса:', tuner_response.status_code)

    def get_remux_parameters(self):
        response = requests.get(
            f'http://{self.device_ip}/cgi-bin/mux.cgi',
            headers=self.headers
        )
        response.encoding = 'utf-8'

        soup = BeautifulSoup(response.text, 'html.parser')

        tuner1_data = soup.find('div', {'id': 'Tuner1_out_value'})
        tuner1_raw = tuner1_data.find_all('div', {'class': 'tree_3'})

        self.tuner1_outputs = [
            div.text.strip().split('\xa0')[-1]
            for div in tuner1_raw
        ]

        tuner2_data = soup.find('div', {'id': 'Tuner2_out_value'})
        tuner2_raw = tuner2_data.find_all('div', {'class': 'tree_3'})

        self.tuner2_outputs = [
            div.text.strip().split('\xa0')[-1]
            for div in tuner2_raw
        ]

        tuner3_data = soup.find('div', {'id': 'Tuner3_out_value'})
        tuner3_raw = tuner3_data.find_all('div', {'class': 'tree_3'})

        self.tuner3_outputs = [
            div.text.strip().split('\xa0')[-1]
            for div in tuner3_raw
        ]

        tuner4_data = soup.find('div', {'id': 'Tuner3_out_value'})
        tuner4_raw = tuner4_data.find_all('div', {'class': 'tree_3'})

        self.tuner4_outputs = [
            div.text.strip().split('\xa0')[-1]
            for div in tuner4_raw
        ]

        # ip_data = soup.find('div', {'id': 'IPIN_out_value'})
        # ip_raw = ip_data.find_all('div', {'class': 'tree_3'})

        # self.ip_outputs = [
        #     div.text.strip().split('\xa0')[-1]
        #     for div in ip_raw
        # ]


    def check_ip_type(self):
        response = requests.get(
            f'http://{self.device_ip}/cgi-bin/ip.cgi',
            headers=self.headers
        )
        soup = BeautifulSoup(response.text, 'html.parser')
        ip_type = soup.find(
            'tr',
            {'class': 'content_title'}
        )
        ip_type = ip_type.text.strip()
        return ip_type


    def get_ip_output(self):
        ip_parts = []
        response = requests.get(
            f'http://{self.device_ip}/cgi-bin/ip_dvb.cgi',
            headers=self.headers
        )
        soup = BeautifulSoup(response.text, 'html.parser')

        for i in range(1, 5):
            input_name = f'dvb_ip{i}'
            input_tag = soup.find('input', {'name': input_name})
            if input_tag:
                ip_parts.append(input_tag['value'])
            else:
                break
        ip_address = '.'.join(ip_parts)
        port = soup.find(
            'input',
            {'name': 'ip_out_multicast_port'}
        ).get('value')
        self.output_ip = f'{ip_address}:{port}'

    def get_ip_input(self):
        ip_parts = []
        response = requests.get(
            f'http://{self.device_ip}/cgi-bin/ip.cgi',
            headers=self.headers
        )
        soup = BeautifulSoup(response.text, 'html.parser')

        for i in range(0, 4):
            input_name = f'multicast_ip{i}'
            input_tag = soup.find('input', {'name': input_name})
            if input_tag:
                ip_parts.append(input_tag['value'])
            else:
                break
        ip_address = '.'.join(ip_parts)
        port = soup.find(
            'input',
            {'name': 'ip_in_multicast_port'}
        ).get('value')
        self.input_ip = f'{ip_address}:{port}'

    def get_ip_params(self):
        ip_type = self.check_ip_type()
        if ip_type == 'TS/IP In':
            self.get_ip_input()
        elif ip_type == 'TS/IP Out':
            self.get_ip_output()

    def get_all_parameters(self):
        self.get_tuner1_parameters()
        self.get_tuner2_parameters()
        self.get_tuner3_parameters()
        self.get_tuner4_parameters()
        self.get_remux_parameters()
        # self.get_decoder()
        # self.get_ip_params()

    def export_params_to_excel(self, name):
        output_dir = os.path.join(os.getcwd(), "excel_output")
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        file_path = os.path.join(
            output_dir,
            f"{self.location} {name} PBI {datetime.date.today()}.xlsx"
        )
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
        worksheet = workbook.active
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        worksheet.column_dimensions['A'].width = 20

        worksheet['A1'] = f'Конфигурация устройства {self.device_ip} ({self.location}) {datetime.date.today()}'
        worksheet['A1'].font = Font(bold=True)
        row = worksheet.max_row + 2

        worksheet.cell(row=row, column=1, value='Параметры Tuner1').font = Font(bold=True)
        row = worksheet.max_row + 1

        worksheet.cell(row=row, column=1, value='LNB').border = border
        worksheet.cell(row=row, column=2, value=self.tuner1_lnb).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='Satellite Frequency').border = border
        worksheet.cell(row=row, column=2, value=self.tuner1_satellite_frequency).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='Symbol Rate').border = border
        worksheet.cell(row=row, column=2, value=self.tuner1_symbol_rate).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='LNB Voltage').border = border
        worksheet.cell(row=row, column=2, value=self.tuner1_lnb_voltage).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='LNB 22KHz').border = border
        worksheet.cell(row=row, column=2, value=self.tuner1_lnb_22khz).border = border
        row += 2

        worksheet.cell(row=row, column=1, value='Параметры Tuner2').font = Font(bold=True)
        row = worksheet.max_row + 1

        worksheet.cell(row=row, column=1, value='LNB').border = border
        worksheet.cell(row=row, column=2, value=self.tuner2_lnb).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='Satellite Frequency').border = border
        worksheet.cell(row=row, column=2, value=self.tuner2_satellite_frequency).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='Symbol Rate').border = border
        worksheet.cell(row=row, column=2, value=self.tuner2_symbol_rate).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='LNB Voltage').border = border
        worksheet.cell(row=row, column=2, value=self.tuner2_lnb_voltage).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='LNB 22KHz').border = border
        worksheet.cell(row=row, column=2, value=self.tuner2_lnb_22khz).border = border
        row += 2

        worksheet.cell(row=row, column=1, value='Параметры Remux').font = Font(bold=True)
        row = worksheet.max_row + 1

        worksheet.cell(row=row, column=1, value='Tuner 1 Out').border = border
        worksheet.cell(row=row, column=2, value=", ".join(str(x) for x in self.tuner1_outputs)).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='Tuner 2 Out').border = border
        worksheet.cell(row=row, column=2, value=", ".join(str(x) for x in self.tuner2_outputs)).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='Мультикаст').font = Font(bold=True)
        row += 1

        if hasattr(self, 'output_ip'):
            worksheet.cell(row=row, column=1, value='IP multicast OUT').border = border
            worksheet.cell(row=row, column=2, value=str(self.output_ip)).border = border
            row += 2

        if hasattr(self, 'input_ip'):
            worksheet.cell(row=row, column=1, value='IP multicast IN').border = border
            worksheet.cell(row=row, column=2, value=str(self.input_ip)).border = border
            row += 2

        workbook.save(file_path)
