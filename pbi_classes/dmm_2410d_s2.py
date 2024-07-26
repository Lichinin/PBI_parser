import base64
import datetime
import os
import time

import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

from selectors.selectors import Selectors as SL


class DMM2410D:
    def __init__(self, device_ip, username, password, location):
        self.device_ip = device_ip
        self.username = username
        self.password = password
        self.location = location
        self.headers = {
            'Authorization': 'Basic ' + base64.b64encode(f'{username}:{password}'.encode('utf-8')).decode('utf-8')
        }

    def check_connection(self):
        tuner_url = f'http://{self.username}:{self.password}@{self.device_ip}'
        response = requests.get(tuner_url)
        return response.status_code

    def get_tuner_parameters(self, tuner_number):
        tuner_url = f'http://{self.username}:{self.password}@{self.device_ip}/tuner{tuner_number}.html'

        chrome_options = Options()
        chrome_options.add_argument("--headless")
        driver = webdriver.Chrome(options=chrome_options)
        driver.get(tuner_url)
        time.sleep(1)

        lnb_freq_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(SL.LnbFreq)
        )
        attribute_name = f"tuner{tuner_number}_lnb"
        setattr(self, attribute_name, lnb_freq_element.get_attribute("value"))

        tuner1_satellite_frequency = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(SL.SateFreq)
        )
        attribute_name = f"tuner{tuner_number}_satellite_frequency"
        setattr(
            self,
            attribute_name,
            tuner1_satellite_frequency.get_attribute("value")
        )

        tuner1_symbol_rate = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(SL.SateSr)
        )
        attribute_name = f"tuner{tuner_number}_symbol_rate"
        setattr(
            self,
            attribute_name,
            tuner1_symbol_rate.get_attribute("value")
        )

        tuner_response = requests.get(tuner_url)
        soup = BeautifulSoup(tuner_response.text, 'html.parser')

        lnb_voltage_raw = soup.find('select', {'id': 'lnbVol'})
        tuner_lnb_voltage = (((lnb_voltage_raw.find(
            'option',
            {'selected': True})
        ).text).split('\n'))[0]
        attribute_name = f"tuner{tuner_number}_lnb_voltage"
        setattr(self, attribute_name, tuner_lnb_voltage)

        lnb_22khz_raw = soup.find('select', {'id': 'lnb22k'})
        tuner_lnb_22khz = (((lnb_22khz_raw.find(
            'option',
            {'selected': True})).text
        ).split('\n'))[0]
        attribute_name = f"tuner{tuner_number}_lnb_22khz"
        setattr(self, attribute_name, tuner_lnb_22khz)

        driver.quit()

    def get_remux_parameters(self):
        username = 'admin'
        password = 'CrnRjhf'
        tuner_url = f'http://{username}:{password}@{self.device_ip}/mux.html'

        chrome_options = Options()
        chrome_options.add_argument("--headless")
        driver = webdriver.Chrome(options=chrome_options)
        driver.get(tuner_url)
        time.sleep(1)

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(SL.Encoder1_btn)
        ).click()
        raw_out = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located(SL.Encoder1))
        self.tuner1_outputs = raw_out[3].text.split()

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(SL.Encoder2_btn)
        ).click()
        raw_out = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located(SL.Encoder2))
        self.tuner2_outputs = raw_out[3].text.split()

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(SL.Encoder3_btn)
        ).click()
        raw_out = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located(SL.Encoder3))
        self.tuner3_outputs = raw_out[3].text.split()

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(SL.Encoder4_btn)
        ).click()
        raw_out = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located(SL.Encoder4))
        self.tuner4_outputs = raw_out[3].text.strip().split('    ')
        driver.quit()

    def get_all_parameters(self):
        if self.check_connection() != 200:
            raise ConnectionError('Нет связи с устройством')
        for tuner in range(1, 5):
            self.get_tuner_parameters(tuner)
        self.get_remux_parameters()

    def export_params_to_excel(self, name):
        output_dir = os.path.join(os.getcwd(), "excel_output")
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        file_path = os.path.join(
            output_dir,
            f"{self.location} {name} PBI {self.__class__.__name__} {datetime.date.today()}.xlsx"
        )
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
        worksheet = workbook.active
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        worksheet.column_dimensions['A'].width = 20

        worksheet['A1'] = f'Конфигурация устройства {self.device_ip} ({self.__class__.__name__}) ({self.location}) {datetime.date.today()}'
        worksheet['A1'].font = Font(bold=True)
        row = worksheet.max_row + 2

        worksheet.cell(
            row=row,
            column=1,
            value='Параметры Tuner1'
        ).font = Font(bold=True)
        row = worksheet.max_row + 1

        worksheet.cell(row=row, column=1, value='LNB').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner1_lnb
        ).border = border
        row += 1

        worksheet.cell(
            row=row,
            column=1,
            value='Satellite Frequency'
        ).border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner1_satellite_frequency
        ).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='Symbol Rate').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner1_symbol_rate
        ).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='LNB Voltage').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner1_lnb_voltage
        ).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='LNB 22KHz').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner1_lnb_22khz
        ).border = border
        row += 2

        worksheet.cell(
            row=row,
            column=1,
            value='Параметры Tuner2'
        ).font = Font(bold=True)
        row = worksheet.max_row + 1

        worksheet.cell(row=row, column=1, value='LNB').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner2_lnb
        ).border = border
        row += 1

        worksheet.cell(
            row=row,
            column=1,
            value='Satellite Frequency'
        ).border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner2_satellite_frequency
        ).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='Symbol Rate').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner2_symbol_rate
        ).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='LNB Voltage').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner2_lnb_voltage
        ).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='LNB 22KHz').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner2_lnb_22khz
        ).border = border
        row += 2

        worksheet.cell(
            row=row,
            column=1,
            value='Параметры Tuner3'
        ).font = Font(bold=True)
        row = worksheet.max_row + 1

        worksheet.cell(row=row, column=1, value='LNB').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner3_lnb
            ).border = border
        row += 1

        worksheet.cell(
            row=row,
            column=1,
            value='Satellite Frequency'
        ).border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner3_satellite_frequency
        ).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='Symbol Rate').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner3_symbol_rate
        ).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='LNB Voltage').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner3_lnb_voltage
        ).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='LNB 22KHz').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner3_lnb_22khz
        ).border = border
        row += 2

        worksheet.cell(
            row=row,
            column=1,
            value='Параметры Tuner4'
        ).font = Font(bold=True)
        row = worksheet.max_row + 1

        worksheet.cell(row=row, column=1, value='LNB').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner4_lnb
        ).border = border
        row += 1

        worksheet.cell(
            row=row,
            column=1,
            value='Satellite Frequency'
        ).border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner4_satellite_frequency
        ).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='Symbol Rate').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner4_symbol_rate
        ).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='LNB Voltage').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner4_lnb_voltage
        ).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='LNB 22KHz').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=self.tuner4_lnb_22khz
        ).border = border
        row += 2

        worksheet.cell(
            row=row,
            column=1,
            value='Параметры Remux'
        ).font = Font(bold=True)
        row = worksheet.max_row + 1

        worksheet.cell(row=row, column=1, value='Tuner 1 Out').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=", ".join(str(x) for x in self.tuner1_outputs)
        ).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='Tuner 2 Out').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=", ".join(str(x) for x in self.tuner2_outputs)
        ).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='Tuner 3 Out').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=", ".join(str(x) for x in self.tuner3_outputs)
        ).border = border
        row += 1

        worksheet.cell(row=row, column=1, value='Tuner 4 Out').border = border
        worksheet.cell(
            row=row,
            column=2,
            value=", ".join(str(x) for x in self.tuner4_outputs)
        ).border = border
        row += 1

        workbook.save(file_path)
